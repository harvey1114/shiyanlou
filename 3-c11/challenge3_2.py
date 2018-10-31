from openpyxl import load_workbook 
from openpyxl import Workbook 
from datetime import datetime 



def combine():
    wb = Workbook()
    path = 'courses.xlsx'
    wb = load_workbook(path)
    sheet_names = wb.sheetnames
    if 'combine' in sheet_names:
        sheet_names.remove('combine')
        wb.remove(wb['combine'])
    ws = wb.active
    ws = wb.create_sheet('combine',index=len(sheet_names))
    for sheet_name in sheet_names:
        sheet = wb[sheet_name]
        for i,row in enumerate(sheet.rows):
            if i==0 and len(list(ws.rows))>0:
                continue
            row_list= []
            for col in range(0,sheet.max_column):
                row_list.append(row[col].value)
            ws.append(row_list)

    wb.save(path)

def split():

    wb = Workbook()
    path = 'courses.xlsx'
    wb = load_workbook(path)
    sheet_names = wb.sheetnames
    for i,column in enumerate(wb['combine'].columns):
        if column[0].value == '创建时间':
            col_num = i
    year_set = set()
    for d in list(wb['combine'].columns)[col_num][1:]:
        year_set.add(d.value.year)
    
    year_dict={}
    for year in year_set:
        year_dict[year]=[]
    for i,row in enumerate(wb['combine'].rows):
        row_list= []
        for col in range(0,wb['combine'].max_column):
            
            row_list.append(row[col].value)
        if i == 0:
            first_row = row_list
        else:
            year_dict[row_list[col_num].year].append(row_list)

    for key,value in year_dict.items():
        wb = Workbook()
        ws = wb.create_sheet(str(key),index=0)
        ws.append(first_row)
        for item in value:
            ws.append(item)
        wb.save(str(key)+'.xlsx')   



if __name__ == '__main__':
    combine()
    split()
